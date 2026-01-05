"""
Transformador ETL mejorado con soporte para CSV
Puede procesar tanto archivos Excel como CSV exportados de Google Sheets
"""

import openpyxl
import pandas as pd
import json
import argparse
from pathlib import Path
from typing import Dict, List, Any


class TriageDataTransformer:
    """
    Transformador flexible que soporta múltiples formatos de entrada
    """
    
    def __init__(self, input_path: str):
        self.input_path = Path(input_path)
        self.knowledge_base = []
        
        # Detectar tipo de archivo
        self.file_type = self.input_path.suffix.lower()
        
    def transform(self) -> List[Dict[str, Any]]:
        """Transforma el archivo según su tipo"""
        
        if self.file_type in ['.xlsx', '.xls']:
            return self._transform_excel()
        elif self.file_type == '.csv':
            return self._transform_csv()
        else:
            raise ValueError(f"Formato no soportado: {self.file_type}")
    
    def _transform_excel(self) -> List[Dict[str, Any]]:
        """Transforma archivo Excel"""
        print(f"📊 Procesando archivo Excel: {self.input_path.name}")
        
        workbook = openpyxl.load_workbook(self.input_path, data_only=True)
        
        for sheet_name in workbook.sheetnames:
            print(f"  📄 Procesando hoja: {sheet_name}")
            
            sheet = workbook[sheet_name]
            sintoma_data = self._extract_sheet_data(sheet, sheet_name)
            
            if sintoma_data:
                self.knowledge_base.append(sintoma_data)
                print(f"    ✅ {len(sintoma_data['preguntas_obligatorias'])} preguntas, "
                      f"{len(sintoma_data['reglas_clasificacion'])} reglas")
        
        return self.knowledge_base
    
    def _transform_csv(self) -> List[Dict[str, Any]]:
        """Transforma archivo CSV (formato aplanado)"""
        print(f"📊 Procesando archivo CSV: {self.input_path.name}")
        
        df = pd.read_csv(self.input_path)
        
        # Agrupar por síntoma
        sintomas = df['sintoma_raiz'].unique()
        
        for sintoma in sintomas:
            sintoma_df = df[df['sintoma_raiz'] == sintoma]
            sintoma_data = self._extract_csv_data(sintoma_df, sintoma)
            
            if sintoma_data:
                self.knowledge_base.append(sintoma_data)
                print(f"  ✅ {sintoma}: {len(sintoma_data['reglas_clasificacion'])} reglas")
        
        return self.knowledge_base
    
    def _extract_sheet_data(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Extrae datos de una hoja de Excel"""
        
        sintoma_data = {
            "sintoma_raiz": sheet_name.lower().strip(),
            "preguntas_obligatorias": [],
            "recomendaciones": [],
            "reglas_clasificacion": []
        }
        
        current_section = None
        
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if not any(row):
                continue
            
            first_cell = str(row[0]).strip().upper() if row[0] else ""
            
            # Detectar secciones
            if "PREGUNTA" in first_cell:
                current_section = "preguntas"
                continue
            elif "RECOMENDACION" in first_cell:
                current_section = "recomendaciones"
                continue
            elif any(code in first_cell for code in ["CODIGO", "CÓDIGO", "CLASIFICACION", "D1", "D2", "D7", "D3"]):
                current_section = "clasificacion"
                # No continue aquí, porque la primera fila puede contener datos
            
            # Procesar según sección
            if current_section == "preguntas":
                pregunta = self._parse_pregunta(row)
                if pregunta:
                    sintoma_data["preguntas_obligatorias"].append(pregunta)
            
            elif current_section == "recomendaciones":
                recomendacion = self._parse_recomendacion(row)
                if recomendacion:
                    sintoma_data["recomendaciones"].append(recomendacion)
            
            elif current_section == "clasificacion":
                regla = self._parse_clasificacion(row)
                if regla:
                    sintoma_data["reglas_clasificacion"].append(regla)
        
        return sintoma_data
    
    def _extract_csv_data(self, df: pd.DataFrame, sintoma: str) -> Dict[str, Any]:
        """Extrae datos de un DataFrame CSV"""
        
        sintoma_data = {
            "sintoma_raiz": sintoma.lower().strip(),
            "preguntas_obligatorias": [],
            "recomendaciones": [],
            "reglas_clasificacion": []
        }
        
        for _, row in df.iterrows():
            # Extraer pregunta si existe
            if pd.notna(row.get('pregunta')):
                pregunta = {
                    "id": f"q_{hash(row['pregunta']) % 10000}",
                    "pregunta": str(row['pregunta']),
                    "tipo_respuesta": str(row.get('tipo_respuesta', 'si_no')),
                    "peso": 1.0
                }
                sintoma_data["preguntas_obligatorias"].append(pregunta)
            
            # Extraer regla de clasificación
            if pd.notna(row.get('codigo_triage')):
                regla = {
                    "codigo_triage": str(row['codigo_triage']),
                    "condiciones": {
                        "pregunta": str(row.get('condicion_pregunta', '')),
                        "respuesta_esperada": str(row.get('respuesta_esperada', ''))
                    },
                    "instruccion_atencion": str(row.get('instruccion', '')),
                    "posibles_causas": str(row.get('causas', '')).split(',') if pd.notna(row.get('causas')) else []
                }
                sintoma_data["reglas_clasificacion"].append(regla)
        
        return sintoma_data
    
    def _parse_pregunta(self, row: tuple) -> Dict[str, Any] | None:
        """Parsea una fila de pregunta"""
        if not row[0]:
            return None
        
        pregunta_texto = str(row[0]).strip()
        
        # Ignorar filas de encabezado
        if pregunta_texto.upper() in ["PREGUNTA", "PREGUNTAS OBLIGATORIAS"]:
            return None
        
        tipo_respuesta = "si_no"
        if len(row) > 1 and row[1]:
            tipo_str = str(row[1]).strip().lower()
            if tipo_str in ["numero", "número", "valor"]:
                tipo_respuesta = "valor"
            elif tipo_str in ["multiple", "opcion"]:
                tipo_respuesta = "multiple"
        
        return {
            "id": f"q_{hash(pregunta_texto) % 10000}",
            "pregunta": pregunta_texto,
            "tipo_respuesta": tipo_respuesta,
            "peso": 1.0
        }
    
    def _parse_recomendacion(self, row: tuple) -> str | None:
        """Parsea una fila de recomendación"""
        if not row[0]:
            return None
        
        texto = str(row[0]).strip()
        
        # Ignorar encabezados
        if texto.upper() in ["RECOMENDACION", "RECOMENDACIONES"]:
            return None
        
        return texto
    
    def _parse_clasificacion(self, row: tuple) -> Dict[str, Any] | None:
        """Parsea una regla de clasificación"""
        if not row[0]:
            return None
        
        primera_celda = str(row[0]).strip().upper()
        
        # Detectar código
        codigo = None
        if "D1" in primera_celda or "EMERGENCIA" in primera_celda or "01" in primera_celda:
            codigo = "D1"
        elif "D2" in primera_celda or ("URGENCIA" in primera_celda and "BAJA" not in primera_celda) or "02" in primera_celda:
            codigo = "D2"
        elif "D7" in primera_celda or "BAJA COMPLEJIDAD" in primera_celda or "07" in primera_celda:
            codigo = "D7"
        elif "D3" in primera_celda or "CONSULTA" in primera_celda or "03" in primera_celda:
            codigo = "D3"
        
        if not codigo:
            return None
        
        # Extraer otros campos
        condicion = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        respuesta = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        instruccion = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        causas = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        
        return {
            "codigo_triage": codigo,
            "condiciones": {
                "pregunta": condicion,
                "respuesta_esperada": respuesta
            },
            "instruccion_atencion": instruccion,
            "posibles_causas": [c.strip() for c in causas.split(",") if c.strip()]
        }
    
    def save_to_json(self, output_path: str):
        """Guarda la base de conocimiento en JSON"""
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(self.knowledge_base, f, ensure_ascii=False, indent=2)
        
        print(f"\n✅ Base de conocimiento guardada en: {output_path}")
        print(f"📊 Total de síntomas procesados: {len(self.knowledge_base)}")
        
        # Estadísticas
        total_preguntas = sum(len(s['preguntas_obligatorias']) for s in self.knowledge_base)
        total_reglas = sum(len(s['reglas_clasificacion']) for s in self.knowledge_base)
        
        print(f"📝 Total de preguntas: {total_preguntas}")
        print(f"⚕️  Total de reglas de clasificación: {total_reglas}")


def main():
    parser = argparse.ArgumentParser(description="Orion Core - Transformador ETL (Excel/CSV → JSON)")
    parser.add_argument("--input", "-i", required=True, help="Archivo de entrada (.xlsx, .xls, .csv)")
    parser.add_argument("--output", "-o", default="data/triage_knowledge_base.json", 
                        help="Archivo JSON de salida")
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("🚀 ORION CORE - TRANSFORMADOR ETL")
    print("=" * 60 + "\n")
    
    try:
        transformer = TriageDataTransformer(args.input)
        transformer.transform()
        transformer.save_to_json(args.output)
        
        print("\n✨ Transformación completada exitosamente")
        
    except Exception as e:
        print(f"\n❌ Error durante la transformación: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
